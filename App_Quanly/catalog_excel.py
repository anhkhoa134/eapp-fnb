"""
Import / export catalog (Category, Product, ProductUnit, Topping, ProductTopping) từ Excel.
"""
from __future__ import annotations

import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction

from App_Catalog.models import Category, Product, ProductTopping, ProductUnit, StoreCategory, StoreProduct, Topping
from App_Tenant.models import Store, Tenant

SHEET_HUONG_DAN = 'Huong_dan'
SHEET_DANH_MUC = 'Danh_muc'
SHEET_SAN_PHAM = 'San_pham'
SHEET_DON_VI = 'Don_vi'
SHEET_TOPPING = 'Topping'
SHEET_SAN_PHAM_TOPPING = 'San_pham_Topping'

MAX_UPLOAD_BYTES = 5 * 1024 * 1024


def _norm_header(val: Any) -> str:
    if val is None:
        return ''
    return str(val).strip().lower().replace(' ', '_')


def _row_is_empty(row: tuple[Any, ...]) -> bool:
    return all(v is None or (isinstance(v, str) and not str(v).strip()) for v in row)


def _iter_sheet_dicts(ws) -> list[tuple[int, dict[str, Any]]]:
    from openpyxl.worksheet.worksheet import Worksheet

    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    out: list[tuple[int, dict[str, Any]]] = []
    for excel_row_idx, row in enumerate(rows[1:], start=2):
        if not row or _row_is_empty(row):
            continue
        d: dict[str, Any] = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            d[key] = row[i] if i < len(row) else None
        out.append((excel_row_idx, d))
    return out


def parse_bool_cell(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) != 0
    s = str(value).strip().lower()
    if not s:
        return default
    if s in ('0', 'false', 'no', 'off', 'tắt', 'tat', 'không', 'khong'):
        return False
    if s in ('1', 'true', 'yes', 'on', 'có', 'co', 'bật', 'bat'):
        return True
    return default


def parse_decimal_cell(value: Any) -> Decimal:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        raise ValueError('Giá trống')
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace(' ', '')
    if not s:
        raise ValueError('Giá trống')
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        parts = s.split(',')
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
            int_part = parts[0].replace('.', '')
            s = f'{int_part}.{parts[1]}'
        else:
            s = s.replace(',', '')
    else:
        if re.fullmatch(r'\d{1,3}(\.\d{3})+', s):
            s = s.replace('.', '')
        elif s.count('.') == 1 and s.split('.')[-1].isdigit() and len(s.split('.')[-1]) <= 2:
            pass
        elif s.count('.') > 1:
            s = s.replace('.', '')
    return Decimal(s)


def parse_int_cell(value: Any, default: int = 0) -> int:
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip()
    try:
        return int(Decimal(s))
    except (InvalidOperation, ValueError):
        return int(float(s))


def resolve_store_ids(tenant: Tenant, raw: Any, stores: list[Store]) -> set[int]:
    all_ids = {s.id for s in stores}
    if raw is None:
        return all_ids
    s = str(raw).strip()
    if not s or s == '*':
        return all_ids
    parts = re.split(r'[,;]', s)
    chosen: set[int] = set()
    for p in parts:
        name = p.strip()
        if not name:
            continue
        if name == '*':
            return all_ids
        lower = name.lower()
        matches = [st for st in stores if st.name.strip().lower() == lower]
        if len(matches) != 1:
            raise ValueError(f'Cửa hàng không xác định: "{name}"')
        chosen.add(matches[0].id)
    return chosen if chosen else all_ids


def sync_category_store_links(category: Category, selected_store_ids: set[int]) -> None:
    all_store_ids = list(Store.objects.filter(tenant=category.tenant).values_list('id', flat=True))
    for store_id in all_store_ids:
        link, _ = StoreCategory.objects.get_or_create(store_id=store_id, category=category)
        link.is_visible = store_id in selected_store_ids
        link.save(update_fields=['is_visible', 'updated_at'])


def sync_product_store_links(product: Product, selected_store_ids: set[int]) -> None:
    all_store_ids = list(Store.objects.filter(tenant=product.tenant).values_list('id', flat=True))
    for store_id in all_store_ids:
        link, _ = StoreProduct.objects.get_or_create(store_id=store_id, product=product)
        link.is_available = store_id in selected_store_ids
        link.save(update_fields=['is_available', 'updated_at'])


def _categories_by_name(tenant: Tenant, name: str) -> list[Category]:
    return list(Category.objects.filter(tenant=tenant, name=name))


def _toppings_by_name(tenant: Tenant, name: str) -> list[Topping]:
    return list(Topping.objects.filter(tenant=tenant, name=name))


def _product_key(cat_name: str, prod_name: str) -> tuple[str, str]:
    return (cat_name.strip(), prod_name.strip())


def resolve_product_qs(tenant: Tenant, cat_name: str, prod_name: str):
    pn = prod_name.strip()
    cn = cat_name.strip()
    if cn:
        return Product.objects.filter(tenant=tenant, name=pn, category__name=cn)
    qs = Product.objects.filter(tenant=tenant, name=pn)
    return qs


def _style_data_sheet_header(ws, column_widths: list[float]) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_template_workbook():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws0 = wb.active
    ws0.title = SHEET_HUONG_DAN
    lines = [
        'eApp FnB — Hướng dẫn import catalog',
        '',
        'Thứ tự xử lý: Danh_muc → San_pham → Don_vi → Topping → San_pham_Topping.',
        'Cột cua_hang: để trống hoặc * = áp dụng tất cả cửa hàng đang hoạt động; hoặc liệt kê tên cửa hàng, phân cách bởi dấu phẩy hoặc chấm phẩy.',
        'hoat_dong: 1 = bật, 0 = tắt (mặc định bật nếu để trống).',
        'Khóa: Danh mục theo ten_danh_muc; Sản phẩm theo (ten_danh_muc + ten_san_pham), ten_danh_muc có thể để trống nếu tên sản phẩm là duy nhất trong tenant.',
        'Đơn vị / gán topping: có thể tham chiếu sản phẩm sẽ được tạo trong cùng file (sheet San_pham).',
        'Đơn vị: (ten_danh_muc, ten_san_pham, ten_don_vi). Topping: ten_topping. Gán topping: (ten_san_pham, ten_topping) + ten_danh_muc khi cần.',
        '',
        'Các sheet dữ liệu có thêm vài dòng mẫu để tham khảo định dạng; có thể xóa hoặc sửa trước khi import.',
    ]
    for i, line in enumerate(lines, start=1):
        ws0.cell(row=i, column=1, value=line)
    ws0.column_dimensions[get_column_letter(1)].width = 92

    sheet_column_widths: dict[str, list[float]] = {
        SHEET_DANH_MUC: [20, 48, 11, 30],
        SHEET_SAN_PHAM: [16, 24, 42, 36, 11, 30],
        SHEET_DON_VI: [16, 24, 14, 12, 9, 11],
        SHEET_TOPPING: [22, 10, 11],
        SHEET_SAN_PHAM_TOPPING: [16, 24, 20, 12, 9, 11],
    }

    sheets_spec: list[tuple[str, list[str], list[list[Any]]]] = [
        (
            SHEET_DANH_MUC,
            ['ten_danh_muc', 'mo_ta', 'hoat_dong', 'cua_hang'],
            [
                ['Đồ uống', 'Danh mục ví dụ — nước, cà phê', 1, '*'],
                ['Món mặn', 'Danh mục ví dụ — cơm, món chính', 1, '*'],
            ],
        ),
        (
            SHEET_SAN_PHAM,
            ['ten_danh_muc', 'ten_san_pham', 'mo_ta', 'url_hinh', 'hoat_dong', 'cua_hang'],
            [
                ['Đồ uống', 'Trà đá chanh', 'Trà đá với chanh tươi', '', 1, '*'],
                ['Đồ uống', 'Cà phê sữa', 'Pha phin truyền thống', '', 1, '*'],
                ['Món mặn', 'Cơm tấm sườn', 'Sườn nướng, bì, chả', '', 1, '*'],
            ],
        ),
        (
            SHEET_DON_VI,
            ['ten_danh_muc', 'ten_san_pham', 'ten_don_vi', 'gia', 'thu_tu', 'hoat_dong'],
            [
                ['Đồ uống', 'Trà đá chanh', 'Ly', 15000, 0, 1],
                ['Đồ uống', 'Cà phê sữa', 'Ly', 25000, 0, 1],
                ['Món mặn', 'Cơm tấm sườn', 'Phần', 55000, 0, 1],
            ],
        ),
        (
            SHEET_TOPPING,
            ['ten_topping', 'thu_tu', 'hoat_dong'],
            [
                ['Thêm đá', 0, 1],
                ['Thêm sữa', 1, 1],
            ],
        ),
        (
            SHEET_SAN_PHAM_TOPPING,
            ['ten_danh_muc', 'ten_san_pham', 'ten_topping', 'gia_them', 'thu_tu', 'hoat_dong'],
            [
                ['Đồ uống', 'Trà đá chanh', 'Thêm đá', 0, 0, 1],
                ['Đồ uống', 'Cà phê sữa', 'Thêm sữa', 5000, 0, 1],
            ],
        ),
    ]
    for title, headers, sample_rows in sheets_spec:
        nws = wb.create_sheet(title)
        nws.append(headers)
        for row in sample_rows:
            nws.append(row)
        widths = sheet_column_widths.get(title)
        if widths:
            _style_data_sheet_header(nws, widths)
    return wb


def template_workbook_bytes() -> bytes:
    buf = io.BytesIO()
    wb = build_template_workbook()
    wb.save(buf)
    return buf.getvalue()


def import_catalog_from_upload(tenant: Tenant, file_obj) -> dict[str, Any]:
    """
    Validate toàn bộ workbook, sau đó upsert trong một transaction.
    Trả về dict: ok, errors (list str), stats (dict đếm), message (str tóm tắt).
    """
    from openpyxl import load_workbook

    errors: list[str] = []
    stats = {
        'categories_created': 0,
        'categories_updated': 0,
        'products_created': 0,
        'products_updated': 0,
        'units_created': 0,
        'units_updated': 0,
        'toppings_created': 0,
        'toppings_updated': 0,
        'mappings_created': 0,
        'mappings_updated': 0,
    }

    stores = list(Store.objects.filter(tenant=tenant, is_active=True).order_by('name'))
    if not stores:
        return {
            'ok': False,
            'errors': ['Tenant chưa có cửa hàng hoạt động.'],
            'stats': stats,
            'message': '',
        }

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        return {
            'ok': False,
            'errors': [f'Không đọc được file Excel: {exc}'],
            'stats': stats,
            'message': '',
        }

    def sheet_dicts(name: str) -> list[tuple[int, dict[str, Any]]]:
        if name not in wb.sheetnames:
            return []
        return _iter_sheet_dicts(wb[name])

    rows_dm = sheet_dicts(SHEET_DANH_MUC)
    rows_sp = sheet_dicts(SHEET_SAN_PHAM)
    rows_dv = sheet_dicts(SHEET_DON_VI)
    rows_tp = sheet_dicts(SHEET_TOPPING)
    rows_spt = sheet_dicts(SHEET_SAN_PHAM_TOPPING)

    names_in_dm_file: set[str] = set()
    for r, d in rows_dm:
        tn = d.get('ten_danh_muc')
        if tn is None or not str(tn).strip():
            errors.append(f'{SHEET_DANH_MUC} dòng {r}: thiếu ten_danh_muc')
            continue
        name = str(tn).strip()
        if name in names_in_dm_file:
            errors.append(f'{SHEET_DANH_MUC} dòng {r}: trùng ten_danh_muc "{name}" trong file')
        names_in_dm_file.add(name)
        try:
            resolve_store_ids(tenant, d.get('cua_hang'), stores)
        except ValueError as e:
            errors.append(f'{SHEET_DANH_MUC} dòng {r}: {e}')
        amb = _categories_by_name(tenant, name)
        if len(amb) > 1:
            errors.append(f'{SHEET_DANH_MUC} dòng {r}: nhiều danh mục trùng tên "{name}" trong DB')

    seen_products: set[tuple[str, str]] = set()
    for r, d in rows_sp:
        tdm = d.get('ten_danh_muc')
        tsp = d.get('ten_san_pham')
        cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
        if tsp is None or not str(tsp).strip():
            errors.append(f'{SHEET_SAN_PHAM} dòng {r}: thiếu ten_san_pham')
            continue
        prod_name = str(tsp).strip()
        key = _product_key(cat_name, prod_name)
        if key in seen_products:
            errors.append(f'{SHEET_SAN_PHAM} dòng {r}: trùng cặp (ten_danh_muc, ten_san_pham)')
        seen_products.add(key)
        if cat_name:
            if not Category.objects.filter(tenant=tenant, name=cat_name).exists() and cat_name not in names_in_dm_file:
                errors.append(
                    f'{SHEET_SAN_PHAM} dòng {r}: không có danh mục "{cat_name}" (trong DB hoặc sheet {SHEET_DANH_MUC})'
                )
        else:
            qn = resolve_product_qs(tenant, '', prod_name)
            c = qn.count()
            if c > 1:
                errors.append(
                    f'{SHEET_SAN_PHAM} dòng {r}: ten_san_pham "{prod_name}" không gắn danh mục nhưng có {c} sản phẩm trùng tên — cần ten_danh_muc'
                )
        try:
            resolve_store_ids(tenant, d.get('cua_hang'), stores)
        except ValueError as e:
            errors.append(f'{SHEET_SAN_PHAM} dòng {r}: {e}')

    seen_units: set[tuple[str, str, str]] = set()
    for r, d in rows_dv:
        for col in ('ten_san_pham', 'ten_don_vi'):
            if d.get(col) is None or not str(d.get(col)).strip():
                errors.append(f'{SHEET_DON_VI} dòng {r}: thiếu {col}')
                break
        else:
            tdm = d.get('ten_danh_muc')
            cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
            prod_name = str(d.get('ten_san_pham')).strip()
            unit_name = str(d.get('ten_don_vi')).strip()
            uk = (cat_name, prod_name, unit_name)
            if uk in seen_units:
                errors.append(f'{SHEET_DON_VI} dòng {r}: trùng bộ (ten_danh_muc, ten_san_pham, ten_don_vi)')
            seen_units.add(uk)
            qs = resolve_product_qs(tenant, cat_name, prod_name)
            cnt = qs.count()
            pkey = _product_key(cat_name, prod_name)
            if cnt == 0 and pkey not in seen_products:
                errors.append(
                    f'{SHEET_DON_VI} dòng {r}: không tìm thấy sản phẩm ({cat_name or "∅"}, {prod_name}) (DB hoặc sheet {SHEET_SAN_PHAM})'
                )
            elif cnt > 1:
                errors.append(
                    f'{SHEET_DON_VI} dòng {r}: nhiều sản phẩm khớp ({cat_name or "∅"}, {prod_name}) — chỉ định ten_danh_muc'
                )
            try:
                parse_decimal_cell(d.get('gia'))
            except ValueError as e:
                errors.append(f'{SHEET_DON_VI} dòng {r}: gia — {e}')

    seen_top: set[str] = set()
    for r, d in rows_tp:
        tt = d.get('ten_topping')
        if tt is None or not str(tt).strip():
            errors.append(f'{SHEET_TOPPING} dòng {r}: thiếu ten_topping')
            continue
        name = str(tt).strip()
        if name in seen_top:
            errors.append(f'{SHEET_TOPPING} dòng {r}: trùng ten_topping trong file')
        seen_top.add(name)
        amb = _toppings_by_name(tenant, name)
        if len(amb) > 1:
            errors.append(f'{SHEET_TOPPING} dòng {r}: nhiều topping trùng tên "{name}" trong DB')

    seen_map: set[tuple[str, str, str]] = set()
    for r, d in rows_spt:
        tsp = d.get('ten_san_pham')
        ttp = d.get('ten_topping')
        if tsp is None or not str(tsp).strip():
            errors.append(f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: thiếu ten_san_pham')
            continue
        if ttp is None or not str(ttp).strip():
            errors.append(f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: thiếu ten_topping')
            continue
        tdm = d.get('ten_danh_muc')
        cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
        prod_name = str(tsp).strip()
        top_name = str(ttp).strip()
        mk = (cat_name, prod_name, top_name)
        if mk in seen_map:
            errors.append(f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: trùng bộ (ten_danh_muc, ten_san_pham, ten_topping)')
        seen_map.add(mk)
        qs = resolve_product_qs(tenant, cat_name, prod_name)
        cnt = qs.count()
        pkey = _product_key(cat_name, prod_name)
        if cnt == 0 and pkey not in seen_products:
            errors.append(
                f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: không tìm thấy sản phẩm ({cat_name or "∅"}, {prod_name}) (DB hoặc sheet {SHEET_SAN_PHAM})'
            )
        elif cnt > 1:
            errors.append(
                f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: nhiều sản phẩm khớp — chỉ định ten_danh_muc'
            )
        if not Topping.objects.filter(tenant=tenant, name=top_name).exists() and top_name not in seen_top:
            if not Topping.objects.filter(tenant=tenant, name=top_name).exists():
                errors.append(
                    f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: không có topping "{top_name}" (DB hoặc sheet {SHEET_TOPPING})'
                )
        try:
            parse_decimal_cell(d.get('gia_them'))
        except ValueError as e:
            errors.append(f'{SHEET_SAN_PHAM_TOPPING} dòng {r}: gia_them — {e}')

    if errors:
        return {'ok': False, 'errors': errors, 'stats': stats, 'message': ''}

    try:
        with transaction.atomic():
            cat_by_name: dict[str, Category] = {c.name: c for c in Category.objects.filter(tenant=tenant)}

            for _, d in rows_dm:
                name = str(d.get('ten_danh_muc')).strip()
                store_ids = resolve_store_ids(tenant, d.get('cua_hang'), stores)
                desc = str(d.get('mo_ta') or '').strip()
                active = parse_bool_cell(d.get('hoat_dong'), default=True)
                if name in cat_by_name:
                    c = cat_by_name[name]
                    c.description = desc
                    c.is_active = active
                    c.save(update_fields=['description', 'is_active', 'updated_at'])
                    stats['categories_updated'] += 1
                else:
                    c = Category(tenant=tenant, name=name, description=desc, is_active=active)
                    c.save()
                    cat_by_name[name] = c
                    stats['categories_created'] += 1
                sync_category_store_links(c, store_ids)

            prod_by_key: dict[tuple[str, str], Product] = {}
            for p in Product.objects.filter(tenant=tenant).select_related('category'):
                cn = p.category.name if p.category_id else ''
                prod_by_key[_product_key(cn, p.name)] = p

            for _, d in rows_sp:
                tdm = d.get('ten_danh_muc')
                cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
                prod_name = str(d.get('ten_san_pham')).strip()
                store_ids = resolve_store_ids(tenant, d.get('cua_hang'), stores)
                long_d = str(d.get('mo_ta') or '')
                raw_img = d.get('url_hinh')
                img = str(raw_img).strip()[:200] if raw_img is not None and str(raw_img).strip() else ''
                active = parse_bool_cell(d.get('hoat_dong'), default=True)

                if cat_name:
                    key = _product_key(cat_name, prod_name)
                    cat = cat_by_name.get(cat_name)
                else:
                    qs_list = list(Product.objects.filter(tenant=tenant, name=prod_name))
                    if len(qs_list) == 1:
                        pr0 = qs_list[0]
                        cn0 = pr0.category.name if pr0.category_id else ''
                        key = _product_key(cn0, prod_name)
                        cat = pr0.category
                    else:
                        key = _product_key('', prod_name)
                        cat = None

                if key in prod_by_key:
                    pr = prod_by_key[key]
                    if cat_name:
                        pr.category = cat
                    pr.description = long_d
                    pr.image_url = img
                    pr.is_active = active
                    pr.save()
                    stats['products_updated'] += 1
                else:
                    pr = Product(
                        tenant=tenant,
                        category=cat,
                        name=prod_name,
                        short_description='',
                        description=long_d,
                        image_url=img,
                        is_active=active,
                    )
                    pr.save()
                    prod_by_key[key] = pr
                    stats['products_created'] += 1
                sync_product_store_links(pr, store_ids)

            for _, d in rows_dv:
                tdm = d.get('ten_danh_muc')
                cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
                prod_name = str(d.get('ten_san_pham')).strip()
                unit_name = str(d.get('ten_don_vi')).strip()
                if cat_name:
                    pr = prod_by_key.get(_product_key(cat_name, prod_name))
                else:
                    qs_list = list(Product.objects.filter(tenant=tenant, name=prod_name))
                    pr = (
                        qs_list[0]
                        if len(qs_list) == 1
                        else prod_by_key.get(_product_key('', prod_name))
                    )
                if pr is None:
                    raise RuntimeError(f'Lỗi nội bộ: không resolve được sản phẩm ({cat_name!r}, {prod_name!r})')
                key = _product_key(
                    pr.category.name if pr.category_id else '',
                    prod_name,
                )
                prod_by_key[key] = pr
                price = parse_decimal_cell(d.get('gia'))
                order = parse_int_cell(d.get('thu_tu'), 0)
                active = parse_bool_cell(d.get('hoat_dong'), default=True)
                unit = ProductUnit.objects.filter(product=pr, name=unit_name).first()
                if unit:
                    unit.price = price
                    unit.display_order = order
                    unit.is_active = active
                    unit.save(update_fields=['price', 'display_order', 'is_active', 'updated_at'])
                    stats['units_updated'] += 1
                else:
                    ProductUnit.objects.create(
                        product=pr,
                        name=unit_name,
                        price=price,
                        sku='',
                        display_order=order,
                        is_active=active,
                    )
                    stats['units_created'] += 1

            top_by_name: dict[str, Topping] = {t.name: t for t in Topping.objects.filter(tenant=tenant)}

            for _, d in rows_tp:
                name = str(d.get('ten_topping')).strip()
                order = parse_int_cell(d.get('thu_tu'), 0)
                active = parse_bool_cell(d.get('hoat_dong'), default=True)
                if name in top_by_name:
                    t = top_by_name[name]
                    t.display_order = order
                    t.is_active = active
                    t.save(update_fields=['display_order', 'is_active', 'updated_at'])
                    stats['toppings_updated'] += 1
                else:
                    t = Topping(tenant=tenant, name=name, display_order=order, is_active=active)
                    t.save()
                    top_by_name[name] = t
                    stats['toppings_created'] += 1

            for _, d in rows_spt:
                tdm = d.get('ten_danh_muc')
                cat_name = str(tdm).strip() if tdm is not None and str(tdm).strip() else ''
                prod_name = str(d.get('ten_san_pham')).strip()
                top_name = str(d.get('ten_topping')).strip()
                if cat_name:
                    pr = prod_by_key.get(_product_key(cat_name, prod_name))
                else:
                    qs_list = list(Product.objects.filter(tenant=tenant, name=prod_name))
                    pr = (
                        qs_list[0]
                        if len(qs_list) == 1
                        else prod_by_key.get(_product_key('', prod_name))
                    )
                if pr is None:
                    raise RuntimeError(f'Lỗi nội bộ: không resolve được sản phẩm ({cat_name!r}, {prod_name!r})')
                topping = top_by_name[top_name]
                price = parse_decimal_cell(d.get('gia_them'))
                order = parse_int_cell(d.get('thu_tu'), 0)
                active = parse_bool_cell(d.get('hoat_dong'), default=True)
                m = ProductTopping.objects.filter(product=pr, topping=topping).first()
                if m:
                    m.price = price
                    m.display_order = order
                    m.is_active = active
                    m.save(update_fields=['price', 'display_order', 'is_active', 'updated_at'])
                    stats['mappings_updated'] += 1
                else:
                    ProductTopping.objects.create(
                        product=pr,
                        topping=topping,
                        price=price,
                        display_order=order,
                        is_active=active,
                    )
                    stats['mappings_created'] += 1
    except Exception as exc:
        return {
            'ok': False,
            'errors': [f'Lỗi khi ghi DB: {exc}'],
            'stats': stats,
            'message': '',
        }

    parts = [
        f'DM +{stats["categories_created"]}/~{stats["categories_updated"]}',
        f'SP +{stats["products_created"]}/~{stats["products_updated"]}',
        f'ĐV +{stats["units_created"]}/~{stats["units_updated"]}',
        f'TP +{stats["toppings_created"]}/~{stats["toppings_updated"]}',
        f'Gán +{stats["mappings_created"]}/~{stats["mappings_updated"]}',
    ]
    message = 'Import thành công. ' + ', '.join(parts) + ' (tạo mới/cập nhật).'
    return {'ok': True, 'errors': [], 'stats': stats, 'message': message}
