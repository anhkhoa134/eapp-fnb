import io
from datetime import timedelta
from decimal import Decimal

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from App_Accounts.models import User
from App_Catalog.models import Category, Product, ProductTopping, ProductUnit, Topping
from App_Sales.models import DiningTable, Order, OrderItem, OrderItemTopping
from App_Tenant.models import Store, Tenant, UserStoreAccess


class QuanlyPermissionTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Demo', public_slug='demo')
        self.store = Store.objects.create(tenant=self.tenant, name='Store 1', is_default=True)

        self.manager = User.objects.create_user(
            username='manager_demo',
            password='123456',
            tenant=self.tenant,
            role=User.Role.MANAGER,
        )
        self.staff = User.objects.create_user(
            username='staff_demo',
            password='123456',
            tenant=self.tenant,
            role=User.Role.STAFF,
        )
        UserStoreAccess.objects.create(user=self.manager, store=self.store, is_default=True)
        UserStoreAccess.objects.create(user=self.staff, store=self.store, is_default=True)

    def test_manager_can_access_dashboard(self):
        self.client.login(username='manager_demo', password='123456')
        res = self.client.get(reverse('App_Quanly:dashboard'))
        self.assertEqual(res.status_code, 200)
        self.assertEqual(self.client.get(reverse('App_Quanly:orders')).status_code, 200)
        self.assertEqual(self.client.get(reverse('App_Quanly:qr_tables')).status_code, 200)

    def test_staff_cannot_access_dashboard(self):
        self.client.login(username='staff_demo', password='123456')
        res = self.client.get(reverse('App_Quanly:dashboard'))
        self.assertEqual(res.status_code, 403)
        self.assertEqual(self.client.get(reverse('App_Quanly:orders')).status_code, 403)
        self.assertEqual(self.client.get(reverse('App_Quanly:qr_tables')).status_code, 403)

    def test_manager_can_access_topping_crud_pages(self):
        self.client.login(username='manager_demo', password='123456')
        self.assertEqual(self.client.get(reverse('App_Quanly:toppings')).status_code, 200)
        redirect_res = self.client.get(reverse('App_Quanly:product_toppings'))
        self.assertEqual(redirect_res.status_code, 302)
        self.assertIn(reverse('App_Quanly:toppings'), redirect_res.url)

    def test_staff_cannot_access_topping_crud_pages(self):
        self.client.login(username='staff_demo', password='123456')
        self.assertEqual(self.client.get(reverse('App_Quanly:toppings')).status_code, 403)
        self.assertEqual(self.client.get(reverse('App_Quanly:product_toppings')).status_code, 403)

    def test_categories_page_uses_modal_crud_actions(self):
        self.client.login(username='manager_demo', password='123456')
        category = Category.objects.create(tenant=self.tenant, name='Món chính')
        res = self.client.get(reverse('App_Quanly:categories'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('data-bs-target="#createCategoryModal"', html)
        self.assertIn('id="editCategoryModal"', html)
        self.assertIn('id="deleteCategoryModal"', html)
        self.assertIn('class="btn btn-sm btn-outline-primary js-edit-category"', html)
        self.assertNotIn(
            f'href="{reverse("App_Quanly:category_edit", kwargs={"pk": category.id})}"',
            html,
        )

    def test_products_page_uses_modal_crud_actions(self):
        self.client.login(username='manager_demo', password='123456')
        category = Category.objects.create(tenant=self.tenant, name='Món chính')
        product = Product.objects.create(tenant=self.tenant, category=category, name='Phở')
        res = self.client.get(reverse('App_Quanly:products'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('id="createProductModal"', html)
        self.assertIn('id="editProductModal"', html)
        self.assertIn('id="deleteProductModal"', html)
        self.assertIn('id="addUnitModal"', html)
        self.assertNotIn(
            f'href="{reverse("App_Quanly:product_edit", kwargs={"pk": product.id})}"',
            html,
        )

    def test_toppings_page_uses_modal_crud_actions(self):
        self.client.login(username='manager_demo', password='123456')
        topping = Topping.objects.create(tenant=self.tenant, name='Thêm trứng')
        res = self.client.get(reverse('App_Quanly:toppings'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('id="createToppingModal"', html)
        self.assertIn('id="createMappingModal"', html)
        self.assertIn('id="editToppingModal"', html)
        self.assertIn('id="editMappingModal"', html)
        self.assertNotIn(
            f'href="{reverse("App_Quanly:topping_edit", kwargs={"pk": topping.id})}"',
            html,
        )

    def test_qr_tables_page_uses_modal_crud_actions(self):
        self.client.login(username='manager_demo', password='123456')
        table = DiningTable.objects.create(
            tenant=self.tenant,
            store=self.store,
            code='A-01',
            name='Bàn A-01',
            is_active=True,
        )
        res = self.client.get(reverse('App_Quanly:qr_tables'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('id="createTableModal"', html)
        self.assertIn('id="editTableModal"', html)
        self.assertIn('id="deleteTableModal"', html)
        self.assertNotIn(
            f'href="{reverse("App_Quanly:qr_table_edit", kwargs={"pk": table.id})}"',
            html,
        )

    def test_staffs_page_uses_modal_password_reset_action(self):
        self.client.login(username='manager_demo', password='123456')
        res = self.client.get(reverse('App_Quanly:staffs'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('id="createStaffModal"', html)
        self.assertIn('id="resetStaffPasswordModal"', html)
        self.assertNotIn(
            f'href="{reverse("App_Quanly:staff_password_reset", kwargs={"pk": self.staff.id})}"',
            html,
        )

    def test_category_edit_get_redirects_back_to_list(self):
        self.client.login(username='manager_demo', password='123456')
        category = Category.objects.create(tenant=self.tenant, name='Đồ ăn')
        res = self.client.get(reverse('App_Quanly:category_edit', kwargs={'pk': category.id}))
        self.assertEqual(res.status_code, 302)
        self.assertEqual(res.url, reverse('App_Quanly:categories'))

    def test_category_edit_post_updates_from_modal_payload(self):
        self.client.login(username='manager_demo', password='123456')
        category = Category.objects.create(tenant=self.tenant, name='Đồ ăn')
        res = self.client.post(
            reverse('App_Quanly:category_edit', kwargs={'pk': category.id}),
            data={
                'name': 'Đồ ăn nóng',
                'description': 'Món nóng',
                'is_active': 'on',
                'store_ids': [str(self.store.id)],
            },
        )
        self.assertEqual(res.status_code, 302)
        category.refresh_from_db()
        self.assertEqual(category.name, 'Đồ ăn nóng')
        self.assertEqual(category.description, 'Món nóng')

    def test_manager_can_access_staff_management_pages(self):
        self.client.login(username='manager_demo', password='123456')
        self.assertEqual(self.client.get(reverse('App_Quanly:staffs')).status_code, 200)
        self.assertEqual(
            self.client.get(reverse('App_Quanly:staff_password_reset', kwargs={'pk': self.staff.id})).status_code,
            200,
        )

    def test_staff_cannot_access_staff_management_pages(self):
        self.client.login(username='staff_demo', password='123456')
        self.assertEqual(self.client.get(reverse('App_Quanly:staffs')).status_code, 403)
        self.assertEqual(
            self.client.get(reverse('App_Quanly:staff_password_reset', kwargs={'pk': self.staff.id})).status_code,
            403,
        )

    def test_manager_can_reset_staff_password(self):
        self.client.login(username='manager_demo', password='123456')
        res = self.client.post(
            reverse('App_Quanly:staff_password_reset', kwargs={'pk': self.staff.id}),
            data={
                'new_password1': 'Reset@1234',
                'new_password2': 'Reset@1234',
            },
        )
        self.assertEqual(res.status_code, 302)
        self.staff.refresh_from_db()
        self.assertTrue(self.staff.check_password('Reset@1234'))


class QuanlyOrderHistoryTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Tenant Orders', public_slug='tenant-orders')
        self.store_1 = Store.objects.create(tenant=self.tenant, name='Store A', is_default=True)
        self.store_2 = Store.objects.create(tenant=self.tenant, name='Store B', is_default=False)

        self.manager = User.objects.create_user(
            username='manager_orders',
            password='123456',
            tenant=self.tenant,
            role=User.Role.MANAGER,
        )
        UserStoreAccess.objects.create(user=self.manager, store=self.store_1, is_default=True)

        self.cashier_1 = User.objects.create_user(
            username='cashier_a',
            password='123456',
            tenant=self.tenant,
            role=User.Role.STAFF,
        )
        self.cashier_2 = User.objects.create_user(
            username='cashier_b',
            password='123456',
            tenant=self.tenant,
            role=User.Role.STAFF,
        )
        UserStoreAccess.objects.create(user=self.cashier_1, store=self.store_1, is_default=True)
        UserStoreAccess.objects.create(user=self.cashier_2, store=self.store_2, is_default=True)

        category = Category.objects.create(tenant=self.tenant, name='Đồ ăn')
        product = Product.objects.create(tenant=self.tenant, category=category, name='Phở Bò')
        unit = ProductUnit.objects.create(product=product, name='Thường', price=Decimal('90000'))
        topping = Topping.objects.create(tenant=self.tenant, name='Thêm trứng')

        self.order_1 = Order.objects.create(
            tenant=self.tenant,
            store=self.store_1,
            cashier=self.cashier_1,
            payment_method=Order.PaymentMethod.CASH,
            status=Order.Status.COMPLETED,
            subtotal=Decimal('100000'),
            tax_rate=Decimal('0'),
            tax_amount=Decimal('0'),
            total_amount=Decimal('100000'),
            customer_paid=Decimal('120000'),
            change_amount=Decimal('20000'),
        )
        self.order_2 = Order.objects.create(
            tenant=self.tenant,
            store=self.store_2,
            cashier=self.cashier_2,
            payment_method=Order.PaymentMethod.CARD,
            status=Order.Status.CANCELLED,
            subtotal=Decimal('70000'),
            tax_rate=Decimal('0'),
            tax_amount=Decimal('0'),
            total_amount=Decimal('70000'),
            customer_paid=Decimal('70000'),
            change_amount=Decimal('0'),
        )
        Order.objects.filter(pk=self.order_1.pk).update(created_at=timezone.now())
        Order.objects.filter(pk=self.order_2.pk).update(created_at=timezone.now() - timedelta(days=10))
        self.order_1.refresh_from_db()
        self.order_2.refresh_from_db()

        self.order_1.order_code = 'ORD-HISTORY-A'
        self.order_1.save(update_fields=['order_code'])
        self.order_2.order_code = 'ORD-HISTORY-B'
        self.order_2.save(update_fields=['order_code'])

        item = OrderItem.objects.create(
            order=self.order_1,
            product=product,
            unit=unit,
            snapshot_product_name='Phở Bò Kobe',
            snapshot_unit_name='Thường',
            unit_price=Decimal('95000'),
            quantity=1,
            note='Ít hành',
            line_total=Decimal('0'),
        )
        OrderItemTopping.objects.create(
            order_item=item,
            topping=topping,
            snapshot_topping_name='Thêm trứng',
            snapshot_price=Decimal('5000'),
        )

    def test_order_history_filters_work(self):
        self.client.login(username='manager_orders', password='123456')
        res = self.client.get(
            reverse('App_Quanly:orders'),
            {
                'store': str(self.store_1.id),
                'payment_method': Order.PaymentMethod.CASH,
                'status': Order.Status.COMPLETED,
                'cashier': str(self.cashier_1.id),
                'q': 'ORD-HISTORY-A',
                'date_from': (timezone.localdate() - timedelta(days=1)).isoformat(),
                'date_to': timezone.localdate().isoformat(),
            },
        )
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('ORD-HISTORY-A', html)
        self.assertNotIn('ORD-HISTORY-B', html)

    def test_order_history_pagination_keeps_query_params(self):
        self.client.login(username='manager_orders', password='123456')
        for idx in range(1, 23):
            order = Order.objects.create(
                tenant=self.tenant,
                store=self.store_1,
                cashier=self.cashier_1,
                payment_method=Order.PaymentMethod.CASH,
                status=Order.Status.COMPLETED,
                subtotal=Decimal('10000'),
                tax_rate=Decimal('0'),
                tax_amount=Decimal('0'),
                total_amount=Decimal('10000'),
                customer_paid=Decimal('10000'),
                change_amount=Decimal('0'),
            )
            order.order_code = f'ORD-PAGE-{idx:02d}'
            order.save(update_fields=['order_code'])

        res = self.client.get(reverse('App_Quanly:orders'), {'q': 'ORD-PAGE'})
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('?q=ORD-PAGE&page=2', html)

    def test_order_history_renders_item_and_topping_snapshot(self):
        self.client.login(username='manager_orders', password='123456')
        res = self.client.get(reverse('App_Quanly:orders'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn('Phở Bò Kobe', html)
        self.assertIn('Thêm trứng', html)


class QuanlyToppingUnifiedFormTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Tenant Top', public_slug='tenant-top')
        self.store = Store.objects.create(tenant=self.tenant, name='Store Top', is_default=True)
        self.manager = User.objects.create_user(
            username='manager_top',
            password='123456',
            tenant=self.tenant,
            role=User.Role.MANAGER,
        )
        UserStoreAccess.objects.create(user=self.manager, store=self.store, is_default=True)

        category = Category.objects.create(tenant=self.tenant, name='Đồ uống')
        self.product = Product.objects.create(tenant=self.tenant, category=category, name='Cà phê sữa')
        self.topping = Topping.objects.create(tenant=self.tenant, name='Thêm sữa')

    def test_create_topping_form_type_topping_success(self):
        self.client.login(username='manager_top', password='123456')
        res = self.client.post(
            reverse('App_Quanly:toppings'),
            data={
                'form_type': 'topping',
                'topping-name': 'Thêm kem',
                'topping-display_order': '1',
                'topping-is_active': 'on',
            },
        )
        self.assertEqual(res.status_code, 302)
        self.assertTrue(Topping.objects.filter(tenant=self.tenant, name='Thêm kem').exists())

    def test_create_mapping_form_type_mapping_success(self):
        self.client.login(username='manager_top', password='123456')
        res = self.client.post(
            reverse('App_Quanly:toppings'),
            data={
                'form_type': 'mapping',
                'mapping-product': str(self.product.id),
                'mapping-topping': str(self.topping.id),
                'mapping-price': '7000',
                'mapping-display_order': '0',
                'mapping-is_active': 'on',
            },
        )
        self.assertEqual(res.status_code, 302)
        self.assertTrue(ProductTopping.objects.filter(product=self.product, topping=self.topping).exists())

    def test_invalid_mapping_form_shows_error_message(self):
        self.client.login(username='manager_top', password='123456')
        res = self.client.post(
            reverse('App_Quanly:toppings'),
            data={
                'form_type': 'mapping',
                'mapping-price': '5000',
            },
        )
        self.assertEqual(res.status_code, 200)
        self.assertIn('Không thể gán topping cho sản phẩm, vui lòng kiểm tra dữ liệu.', res.content.decode('utf-8'))


class QuanlyQrTableCrudTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='Tenant QR Table', public_slug='tenant-qr-table')
        self.store_1 = Store.objects.create(tenant=self.tenant, name='Store QR 1', is_default=True)
        self.store_2 = Store.objects.create(tenant=self.tenant, name='Store QR 2', is_default=False)
        self.manager = User.objects.create_user(
            username='manager_qr_table',
            password='123456',
            tenant=self.tenant,
            role=User.Role.MANAGER,
        )
        UserStoreAccess.objects.create(user=self.manager, store=self.store_1, is_default=True)

        self.table = DiningTable.objects.create(
            tenant=self.tenant,
            store=self.store_1,
            code='A-01',
            name='Bàn A-01',
            is_active=True,
            display_order=1,
        )

    def test_qr_table_list_and_create(self):
        self.client.login(username='manager_qr_table', password='123456')
        res = self.client.get(reverse('App_Quanly:qr_tables'))
        self.assertEqual(res.status_code, 200)
        self.assertIn('Quản lý QR bàn', res.content.decode('utf-8'))

        post_res = self.client.post(
            reverse('App_Quanly:qr_tables'),
            data={
                'store': str(self.store_2.id),
                'code': 'b-02',
                'name': 'Bàn B-02',
                'display_order': 2,
                'is_active': 'on',
            },
        )
        self.assertEqual(post_res.status_code, 302)
        created = DiningTable.objects.get(tenant=self.tenant, store=self.store_2, code='B-02')
        self.assertEqual(created.name, 'Bàn B-02')

    def test_qr_table_edit(self):
        self.client.login(username='manager_qr_table', password='123456')
        res = self.client.post(
            reverse('App_Quanly:qr_table_edit', kwargs={'pk': self.table.id}),
            data={
                'store': str(self.store_1.id),
                'code': 'a-09',
                'name': 'Bàn VIP 09',
                'display_order': 9,
                'is_active': 'on',
            },
        )
        self.assertEqual(res.status_code, 302)
        self.table.refresh_from_db()
        self.assertEqual(self.table.code, 'A-09')
        self.assertEqual(self.table.name, 'Bàn VIP 09')
        self.assertEqual(self.table.display_order, 9)

    def test_qr_table_reset_token(self):
        self.client.login(username='manager_qr_table', password='123456')
        old_token = self.table.qr_token
        res = self.client.post(reverse('App_Quanly:qr_table_reset_token', kwargs={'pk': self.table.id}))
        self.assertEqual(res.status_code, 302)
        self.table.refresh_from_db()
        self.assertNotEqual(old_token, self.table.qr_token)

    def test_qr_table_delete(self):
        self.client.login(username='manager_qr_table', password='123456')
        res = self.client.post(reverse('App_Quanly:qr_table_delete', kwargs={'pk': self.table.id}))
        self.assertEqual(res.status_code, 302)
        self.assertFalse(DiningTable.objects.filter(pk=self.table.id).exists())

    def test_qr_table_png_download(self):
        self.client.login(username='manager_qr_table', password='123456')
        res = self.client.get(reverse('App_Quanly:qr_table_png', kwargs={'pk': self.table.id}))
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], 'image/png')
        content = b''.join(res.streaming_content)
        self.assertTrue(content.startswith(b'\x89PNG'))

    def test_qr_table_print_pdf_by_store(self):
        self.client.login(username='manager_qr_table', password='123456')
        DiningTable.objects.create(
            tenant=self.tenant,
            store=self.store_1,
            code='A-02',
            name='Bàn A-02',
            is_active=True,
            display_order=2,
        )
        res = self.client.get(reverse('App_Quanly:qr_tables_print_pdf'), {'store': self.store_1.id})
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], 'application/pdf')
        content = b''.join(res.streaming_content)
        self.assertTrue(content.startswith(b'%PDF'))


class CatalogExcelImportTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name='ImpTest', public_slug='imptest')
        self.store = Store.objects.create(tenant=self.tenant, name='Cua 1', is_default=True)
        self.manager = User.objects.create_user(
            username='mgr_imp',
            password='123456',
            tenant=self.tenant,
            role=User.Role.MANAGER,
        )
        UserStoreAccess.objects.create(user=self.manager, store=self.store, is_default=True)
        self.staff = User.objects.create_user(
            username='stf_imp',
            password='123456',
            tenant=self.tenant,
            role=User.Role.STAFF,
        )
        UserStoreAccess.objects.create(user=self.staff, store=self.store, is_default=True)

    def _build_min_workbook(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = 'Huong_dan'
        ws['A1'] = 'test'
        specs = [
            (
                'Danh_muc',
                ['ten_danh_muc', 'mo_ta', 'hoat_dong', 'cua_hang'],
                [['Nước', 'Mô tả', 1, '*']],
            ),
            (
                'San_pham',
                ['ten_danh_muc', 'ten_san_pham', 'mo_ta', 'url_hinh', 'hoat_dong', 'cua_hang'],
                [['Nước', 'Trà đá', '', '', 1, '*']],
            ),
            (
                'Don_vi',
                ['ten_danh_muc', 'ten_san_pham', 'ten_don_vi', 'gia', 'thu_tu', 'hoat_dong'],
                [['Nước', 'Trà đá', 'Ly', '10000', 0, 1]],
            ),
        ]
        for title, headers, rows in specs:
            nws = wb.create_sheet(title)
            nws.append(headers)
            for row in rows:
                nws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_categories_page_shows_excel_import(self):
        self.client.login(username='mgr_imp', password='123456')
        res = self.client.get(reverse('App_Quanly:categories'))
        self.assertEqual(res.status_code, 200)
        html = res.content.decode('utf-8')
        self.assertIn(reverse('App_Quanly:catalog_import_template'), html)
        self.assertIn(reverse('App_Quanly:catalog_import_upload'), html)

    def test_manager_download_template(self):
        self.client.login(username='mgr_imp', password='123456')
        res = self.client.get(reverse('App_Quanly:catalog_import_template'))
        self.assertEqual(res.status_code, 200)
        self.assertIn('spreadsheetml', res['Content-Type'])

    def test_staff_forbidden_template_and_upload(self):
        self.client.login(username='stf_imp', password='123456')
        self.assertEqual(self.client.get(reverse('App_Quanly:catalog_import_template')).status_code, 403)
        buf = self._build_min_workbook()
        res = self.client.post(
            reverse('App_Quanly:catalog_import_upload'),
            {
                'excel_file': SimpleUploadedFile(
                    't.xlsx',
                    buf.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            },
        )
        self.assertEqual(res.status_code, 403)

    def test_manager_import_creates_catalog(self):
        self.client.login(username='mgr_imp', password='123456')
        buf = self._build_min_workbook()
        res = self.client.post(
            reverse('App_Quanly:catalog_import_upload'),
            {
                'excel_file': SimpleUploadedFile(
                    't.xlsx',
                    buf.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            },
        )
        self.assertEqual(res.status_code, 302)
        cat = Category.objects.get(tenant=self.tenant, name='Nước')
        self.assertEqual(cat.description, 'Mô tả')
        p = Product.objects.get(tenant=self.tenant, name='Trà đá')
        self.assertEqual(p.category_id, cat.id)
        u = ProductUnit.objects.get(product=p, name='Ly')
        self.assertEqual(u.price, Decimal('10000'))

    def test_import_duplicate_category_rows_rejected(self):
        from openpyxl import Workbook

        self.client.login(username='mgr_imp', password='123456')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Huong_dan'
        dm = wb.create_sheet('Danh_muc')
        dm.append(['ten_danh_muc', 'mo_ta', 'hoat_dong', 'cua_hang'])
        dm.append(['Trùng', 'a', 1, '*'])
        dm.append(['Trùng', 'b', 1, '*'])
        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()
        res = self.client.post(
            reverse('App_Quanly:catalog_import_upload'),
            {
                'excel_file': SimpleUploadedFile(
                    't.xlsx',
                    raw,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            },
        )
        self.assertEqual(res.status_code, 302)
        self.assertFalse(Category.objects.filter(tenant=self.tenant, name='Trùng').exists())
